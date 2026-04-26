"""Write computed EuroSCORE II scores back to EURO -CR.xlsx.

For each YAML in _emr_raw/:
  1. Update main sheet column J (EURO score) by matching chart number (col D)
  2. Update the per-patient block in '<doctor> Euroscore評估表' sheet:
     fill the inputs and the auto-computed score

Usage: python write_to_excel.py [--doctor 陳常胤] [--file "EURO -CR.xlsx"]
"""
import sys, io, argparse
from pathlib import Path
import openpyxl
import yaml

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from euroscore_ii import calc, cockcroft_gault, renal_category

ROOT = Path(__file__).parent
DEFAULT_FILE = ROOT / "EURO -CR.xlsx"
DEFAULT_DOCTOR = "陳常胤"
SOURCE_SHEET = "2026 (資訊室CAD-3.2)"
COL_CHART = 4
COL_EURO = 10

# Mapping field key -> display value for the per-patient block
RENAL_LABEL = {"normal": "Normal", "cc_50_85": "Moderate (50-85)",
               "cc_le_50": "Severe (<50)", "dialysis": "Dialysis"}
LV_LABEL = {"good": "Good (>50%)", "moderate": "Moderate (31-50%)",
            "poor": "Poor (21-30%)", "very_poor": "Very poor (<20%)"}
PA_LABEL = {"none": "No", "31_55": "Moderate (31-55)", "ge_55": "Severe (>55)"}
URG_LABEL = {"elective": "Elective", "urgent": "Urgent",
             "emergency": "Emergency", "salvage": "Salvage"}
WEIGHT_LABEL = {"isolated_cabg": "Isolated CABG", "single_non_cabg": "Single non-CABG",
                "two": "2", "three_plus": "3+"}


def yn(v):
    return "Y" if v else "N"


def display(p):
    """Build {label: value} dict in the block-row order."""
    cc = cockcroft_gault(p.get("age"), p.get("weight_kg"),
                         p.get("cr_mg_dl"), p.get("female", False))
    cc_str = f"{cc:.1f}" if cc else ""
    renal = p.get("renal")
    if renal != "dialysis" and cc is not None:
        renal = renal_category(cc)
    return {
        "年齡 (歲)": p.get("age"),
        "性別": "女" if p.get("female") else "男",
        "體重 (kg)": p.get("weight_kg"),
        "血清 Cr (mg/dL)": p.get("cr_mg_dl"),
        "Creatinine clearance (ml/min)": cc_str,
        "腎功能分級": RENAL_LABEL.get(renal, renal),
        "NYHA class": p.get("nyha", "I"),
        "CCS angina class 4": yn(p.get("ccs4")),
        "胰島素糖尿病 IDDM": yn(p.get("iddm")),
        "心外動脈病變 ECA": yn(p.get("extracardiac_arteriopathy")),
        "慢性肺病 CPD": yn(p.get("chronic_pulmonary_disease")),
        "行動不便": yn(p.get("poor_mobility")),
        "曾接受心臟手術 Redo": yn(p.get("previous_cardiac_surgery")),
        "Active endocarditis": yn(p.get("active_endocarditis")),
        "Critical preop state": yn(p.get("critical_preop")),
        "LV function": LV_LABEL.get(p.get("lv_function", "good")),
        "Recent MI (90 天內)": yn(p.get("recent_mi")),
        "PA systolic 肺動脈收縮壓": PA_LABEL.get(p.get("pa_systolic", "none")),
        "Urgency": URG_LABEL.get(p.get("urgency", "elective")),
        "Weight of procedure": WEIGHT_LABEL.get(p.get("weight_of_procedure", "isolated_cabg")),
        "Thoracic aorta surgery": yn(p.get("thoracic_aorta")),
    }


def update_main(ws, chart, score):
    """Update column J of the row matching chart number in column D."""
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=COL_CHART).value or "").strip() == chart:
            ws.cell(row=r, column=COL_EURO, value=round(score, 4))
            return r
    return None


def update_assessment_block(ws, chart, score, p):
    """Find the patient block (header contains chart number) and fill values."""
    fields = display(p)
    # Find header row
    header_row = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and chart in str(v):
            header_row = r
            break
    if header_row is None:
        return None
    # Items start 2 rows below header (header, then 項目/值/說明 row, then items)
    item_start = header_row + 2
    filled = 0
    for offset in range(40):  # safety cap
        r = item_start + offset
        label = ws.cell(row=r, column=1).value
        if not label:
            break
        label_str = str(label).strip()
        if "EuroSCORE II 預測死亡率" in label_str:
            ws.cell(row=r, column=2, value=round(score * 100, 2))
            filled += 1
            break
        if label_str in fields:
            ws.cell(row=r, column=2, value=fields[label_str])
            filled += 1
    return filled


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", default=str(DEFAULT_FILE))
    ap.add_argument("--doctor", default=DEFAULT_DOCTOR)
    args = ap.parse_args()

    path = Path(args.file)
    wb = openpyxl.load_workbook(path)
    main_ws = wb[SOURCE_SHEET]
    sheet_name = f"{args.doctor} Euroscore評估表"
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"找不到工作表：{sheet_name}")
    ax_ws = wb[sheet_name]

    rows = []
    for ypath in sorted((ROOT / "_emr_raw").glob("*.yaml")):
        with open(ypath, "r", encoding="utf-8") as f:
            p = yaml.safe_load(f)
        # Apply renal classification (same logic as cheatsheet.py)
        cc_calc = cockcroft_gault(p.get("age"), p.get("weight_kg"),
                                   p.get("cr_mg_dl"), p.get("female", False))
        if p.get("renal") == "dialysis":
            pass  # keep dialysis override
        elif cc_calc is not None:
            p["renal"] = renal_category(cc_calc)
        score = calc(p)
        manual = p.get("manual_score")
        main_row = update_main(main_ws, p["chart"], score)
        block_filled = update_assessment_block(ax_ws, p["chart"], score, p)
        rows.append({
            "chart": p["chart"], "name": p["name"],
            "predicted": score, "manual": manual,
            "main_row": main_row, "block_filled": block_filled,
        })

    wb.save(path)

    print("\n寫入結果：")
    print(f"{'病歷號':<10} {'姓名':<8} {'預測':>7} {'manual':>7} {'diff':>7}  主表列  block")
    for r in rows:
        m = r["manual"]
        diff = (r["predicted"] - m) * 100 if m is not None else None
        diff_str = f"{diff:+.2f}%" if diff is not None else "n/a"
        manual_str = f"{m*100:.2f}%" if m is not None else "n/a"
        pred_str = f"{r['predicted']*100:.2f}%"
        main_str = str(r["main_row"]) if r["main_row"] else "NOT FOUND"
        block_str = f"{r['block_filled']} cells" if r["block_filled"] else "BLOCK NOT FOUND"
        print(f"{r['chart']:<10} {r['name']:<8} {pred_str:>7} {manual_str:>7} {diff_str:>7}  {main_str:>6}  {block_str}")


if __name__ == "__main__":
    main()
