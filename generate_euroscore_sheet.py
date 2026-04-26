"""Generate a per-doctor Euroscore II assessment sheet.

For the given CR doctor, finds rows in the source sheet where:
  - column H (CR均分) matches the doctor name
  - column J (EURO score) is empty
and writes a worksheet "<doctor> Euroscore評估表" with:
  1. A summary table of all assessable patients
  2. One detailed assessment block per patient (18 EuroSCORE II inputs +
     auto-computed score)

After the user fills the input cells, run `recalc_euroscore_sheet.py` to
re-compute and write back the EuroSCORE II values.
"""
import sys
import io
import argparse
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

DEFAULT_FILE = Path(__file__).parent / "EURO -CR.xlsx"
SOURCE_SHEET = "2026 (資訊室CAD-3.2)"

COL_ADMIT = 2
COL_DC = 3
COL_CHART = 4
COL_NAME = 5
COL_DX = 6
COL_VS = 7
COL_CR = 8
COL_EURO = 10

# (label, key, type, options/help)
ITEMS = [
    ("年齡 (歲)", "age", "number", "整數，>=18"),
    ("性別", "female", "choice", "男 / 女"),
    ("體重 (kg)", "weight_kg", "number", "用於計算 CC"),
    ("血清 Cr (mg/dL)", "cr_mg_dl", "number", "用於計算 CC"),
    ("Creatinine clearance (ml/min)", "cc", "auto", "Cockcroft-Gault 自動算"),
    ("腎功能分級", "renal", "choice", "正常(>85) / cc_50_85 / cc_le_50 / 透析"),
    ("NYHA class", "nyha", "choice", "I / II / III / IV"),
    ("CCS angina class 4", "ccs4", "yn", "是 / 否 (CCS4 才算)"),
    ("胰島素糖尿病 IDDM", "iddm", "yn", "是 / 否"),
    ("心外動脈病變 ECA", "extracardiac_arteriopathy", "yn",
     "claudication / 頸動脈>50% / 截肢 / 主動脈或周邊動脈介入"),
    ("慢性肺病 CPD", "chronic_pulmonary_disease", "yn", "長期使用支氣管擴張劑或類固醇"),
    ("行動不便", "poor_mobility", "yn", "嚴重神經/骨骼疾病影響行動"),
    ("曾接受心臟手術 Redo", "previous_cardiac_surgery", "yn",
     "之前曾打開心包進行心臟手術"),
    ("Active endocarditis", "active_endocarditis", "yn", "手術時仍在使用抗生素治療 IE"),
    ("Critical preop state", "critical_preop", "yn",
     "VT/VF/SCD、CPR、術前插管、inotrope、IABP/VAD、急性腎衰其中之一"),
    ("LV function", "lv_function", "choice",
     "good(≥51%) / moderate(31-50%) / poor(21-30%) / very_poor(≤20%)"),
    ("Recent MI (90 天內)", "recent_mi", "yn", "是 / 否"),
    ("PA systolic 肺動脈收縮壓", "pa_systolic", "choice",
     "none(<31) / 31_55 / ge_55"),
    ("Urgency", "urgency", "choice", "elective / urgent / emergency / salvage"),
    ("Weight of procedure", "weight_of_procedure", "choice",
     "isolated_cabg / single_non_cabg / two / three_plus"),
    ("Thoracic aorta surgery", "thoracic_aorta", "yn",
     "升/弓/降主動脈或主動脈根部手術"),
]


HEADER_FILL = PatternFill("solid", fgColor="DCE6F1")
PATIENT_FILL = PatternFill("solid", fgColor="FCE4D6")
SCORE_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fmt_date(v):
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def find_pending(ws, doctor):
    rows = []
    for r in range(2, ws.max_row + 1):
        cr = ws.cell(row=r, column=COL_CR).value
        euro = ws.cell(row=r, column=COL_EURO).value
        if cr is None or doctor not in str(cr):
            continue
        if euro not in (None, ""):
            continue
        rows.append({
            "row": r,
            "chart": str(ws.cell(row=r, column=COL_CHART).value or "").strip(),
            "name": str(ws.cell(row=r, column=COL_NAME).value or "").strip(),
            "admit": fmt_date(ws.cell(row=r, column=COL_ADMIT).value),
            "dc": fmt_date(ws.cell(row=r, column=COL_DC).value),
            "dx": str(ws.cell(row=r, column=COL_DX).value or "").strip(),
            "vs": str(ws.cell(row=r, column=COL_VS).value or "").strip(),
        })
    return rows


def write_summary(ws, rows, start_row=1):
    ws.cell(row=start_row, column=1, value="序").font = Font(bold=True)
    ws.cell(row=start_row, column=2, value="病歷號").font = Font(bold=True)
    ws.cell(row=start_row, column=3, value="姓名").font = Font(bold=True)
    ws.cell(row=start_row, column=4, value="入院").font = Font(bold=True)
    ws.cell(row=start_row, column=5, value="出院").font = Font(bold=True)
    ws.cell(row=start_row, column=6, value="主治").font = Font(bold=True)
    ws.cell(row=start_row, column=7, value="EuroSCORE II %").font = Font(bold=True)
    for c in range(1, 8):
        ws.cell(row=start_row, column=c).fill = HEADER_FILL
        ws.cell(row=start_row, column=c).border = BORDER
    for i, r in enumerate(rows, 1):
        rr = start_row + i
        ws.cell(row=rr, column=1, value=i)
        ws.cell(row=rr, column=2, value=r["chart"])
        ws.cell(row=rr, column=3, value=r["name"])
        ws.cell(row=rr, column=4, value=r["admit"])
        ws.cell(row=rr, column=5, value=r["dc"])
        ws.cell(row=rr, column=6, value=r["vs"])
        for c in range(1, 8):
            ws.cell(row=rr, column=c).border = BORDER
    return start_row + len(rows) + 1


def write_patient_block(ws, patient, start_row):
    """Write a per-patient assessment block. Returns next free row."""
    r = start_row
    title = f"【{patient['name']}】病歷號 {patient['chart']} | 入院 {patient['admit']} → 出院 {patient['dc']}"
    cell = ws.cell(row=r, column=1, value=title)
    cell.font = Font(bold=True, size=12)
    cell.fill = PATIENT_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    if patient.get("dx"):
        c2 = ws.cell(row=r, column=5, value=f"診斷: {patient['dx']}")
        c2.font = Font(italic=True)
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=8)
    r += 1

    ws.cell(row=r, column=1, value="項目").font = Font(bold=True)
    ws.cell(row=r, column=2, value="值").font = Font(bold=True)
    ws.cell(row=r, column=3, value="說明").font = Font(bold=True)
    for c in (1, 2, 3):
        ws.cell(row=r, column=c).fill = HEADER_FILL
    r += 1

    for label, key, kind, helptxt in ITEMS:
        ws.cell(row=r, column=1, value=label)
        if kind == "auto":
            cc_cell = ws.cell(row=r, column=2)
            cc_cell.value = ""
            cc_cell.comment = Comment("自動由體重 / Cr 計算", "system")
        ws.cell(row=r, column=3, value=helptxt).font = Font(color="808080", size=9)
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=2).border = BORDER
        ws.cell(row=r, column=3).border = BORDER
        r += 1

    score_label = ws.cell(row=r, column=1, value="EuroSCORE II 預測死亡率 (%)")
    score_label.font = Font(bold=True, size=11)
    score_label.fill = SCORE_FILL
    score_cell = ws.cell(row=r, column=2)
    score_cell.fill = SCORE_FILL
    score_cell.font = Font(bold=True, size=11)
    score_cell.border = BORDER
    ws.cell(row=r, column=1).border = BORDER
    r += 2
    return r


def write_assessment_sheet(wb, doctor, rows):
    sheet_name = f"{doctor} Euroscore評估表"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 14

    next_row = write_summary(ws, rows, start_row=1)
    next_row += 1

    for p in rows:
        next_row = write_patient_block(ws, p, next_row)

    return sheet_name


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("doctor", help="負責 CR 醫師姓名，例如 陳常胤")
    ap.add_argument("--file", default=str(DEFAULT_FILE))
    ap.add_argument("--sheet", default=SOURCE_SHEET)
    args = ap.parse_args()

    path = Path(args.file)
    wb = openpyxl.load_workbook(path)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"找不到工作表: {args.sheet}\n可用: {wb.sheetnames}")
    ws = wb[args.sheet]

    rows = find_pending(ws, args.doctor)
    print(f"找到 {len(rows)} 筆 {args.doctor} 待評估病人:")
    for r in rows:
        print(f"  列{r['row']}: {r['chart']} {r['name']} {r['admit']} -> {r['dc']}")

    if not rows:
        print("沒有資料可寫入。")
        return

    sheet_name = write_assessment_sheet(wb, args.doctor, rows)
    wb.save(path)
    print(f"\n已寫入工作表「{sheet_name}」到 {path.name}")
    print(f"每位病人有 {len(ITEMS)} 個輸入項目 + 自動算分欄")


if __name__ == "__main__":
    main()
