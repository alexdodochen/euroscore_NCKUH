"""Add an MDCalc-order key-in sheet to EURO -CR.xlsx.

Creates a NEW worksheet 'MDCalc 順序' alongside the existing 陳常胤 sheet.
Each patient block lists fields in the EXACT order of the hospital MDCalc
form (per user screenshot), so the user can copy values straight in.

Columns:
    A: field label (English, matches MDCalc UI)
    B: value (Y/N or category text)
    C: β coefficient contribution
    D: rationale / source

Run: python mdcalc_sheet.py
"""
import sys, io
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import yaml

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from euroscore_ii import calc, cockcroft_gault, renal_category, BETA, age_x

ROOT = Path(__file__).parent
PATH = ROOT / "EURO -CR.xlsx"
SHEET_NAME = "MDCalc 順序"

# (label, value_fn, beta_fn, yaml_key) — in MDCalc UI order
PATIENT_FIELDS = [
    ("Age", lambda p: p.get("age"),
        lambda p: BETA["age"] * age_x(p.get("age")), "age"),
    ("Gender", lambda p: "Female" if p.get("female") else "Male",
        lambda p: BETA["female"] if p.get("female") else 0.0, "female"),
    ("Renal impairment", lambda p: {
            "normal": "Normal (CC>85)",
            "cc_50_85": "Moderate (CC 50-85)",
            "cc_le_50": "Severe (CC<50)",
            "dialysis": "Dialysis",
        }.get(p.get("renal", "normal"), p.get("renal")),
        lambda p: {"dialysis": BETA["dialysis"], "cc_le_50": BETA["cc_le_50"],
                    "cc_50_85": BETA["cc_50_85"]}.get(p.get("renal", "normal"), 0.0),
        "renal"),
    ("Extracardiac arteriopathy", lambda p: "Y" if p.get("extracardiac_arteriopathy") else "N",
        lambda p: BETA["extracardiac_arteriopathy"] if p.get("extracardiac_arteriopathy") else 0.0,
        "extracardiac_arteriopathy"),
    ("Poor mobility", lambda p: "Y" if p.get("poor_mobility") else "N",
        lambda p: BETA["poor_mobility"] if p.get("poor_mobility") else 0.0, "poor_mobility"),
    ("Previous cardiac surgery", lambda p: "Y" if p.get("previous_cardiac_surgery") else "N",
        lambda p: BETA["previous_cardiac_surgery"] if p.get("previous_cardiac_surgery") else 0.0,
        "previous_cardiac_surgery"),
    ("Chronic lung disease", lambda p: "Y" if p.get("chronic_pulmonary_disease") else "N",
        lambda p: BETA["chronic_pulmonary_disease"] if p.get("chronic_pulmonary_disease") else 0.0,
        "chronic_pulmonary_disease"),
    ("Active endocarditis", lambda p: "Y" if p.get("active_endocarditis") else "N",
        lambda p: BETA["active_endocarditis"] if p.get("active_endocarditis") else 0.0,
        "active_endocarditis"),
    ("Critical preoperative state", lambda p: "Y" if p.get("critical_preop") else "N",
        lambda p: BETA["critical_preop"] if p.get("critical_preop") else 0.0, "critical_preop"),
    ("Diabetes on insulin", lambda p: "Y" if p.get("iddm") else "N",
        lambda p: BETA["iddm"] if p.get("iddm") else 0.0, "iddm"),
]

CARDIAC_FIELDS = [
    ("NYHA", lambda p: p.get("nyha", "I"),
        lambda p: {"II": BETA["nyha_ii"], "III": BETA["nyha_iii"], "IV": BETA["nyha_iv"]
                    }.get((p.get("nyha") or "I").upper(), 0.0), "nyha"),
    ("CCS class 4 angina", lambda p: "Y" if p.get("ccs4") else "N",
        lambda p: BETA["ccs4"] if p.get("ccs4") else 0.0, "ccs4"),
    ("LV function", lambda p: {
            "good": "Good (>50%)", "moderate": "Moderate (31-50%)",
            "poor": "Poor (21-30%)", "very_poor": "Very poor (≤20%)",
        }.get(p.get("lv_function", "good"), "Good"),
        lambda p: {"moderate": BETA["lv_moderate"], "poor": BETA["lv_poor"],
                    "very_poor": BETA["lv_very_poor"]}.get(p.get("lv_function", "good"), 0.0),
        "lv_function"),
    ("Recent MI", lambda p: "Y" if p.get("recent_mi") else "N",
        lambda p: BETA["recent_mi"] if p.get("recent_mi") else 0.0, "recent_mi"),
    ("Pulmonary hypertension", lambda p: {
            "none": "No (PASP<31)", "31_55": "Moderate (31-55)", "ge_55": "Severe (>55)",
        }.get(p.get("pa_systolic", "none"), "No"),
        lambda p: {"31_55": BETA["pa_31_55"], "ge_55": BETA["pa_ge_55"]
                    }.get(p.get("pa_systolic", "none"), 0.0), "pa_systolic"),
]

OPERATION_FIELDS = [
    ("Urgency", lambda p: {"elective": "Elective", "urgent": "Urgent",
                            "emergency": "Emergency", "salvage": "Salvage"
                            }.get(p.get("urgency", "elective"), "Elective"),
        lambda p: {"urgent": BETA["urgent"], "emergency": BETA["emergency"],
                    "salvage": BETA["salvage"]}.get(p.get("urgency", "elective"), 0.0),
        "urgency"),
    ("Weight of the intervention", lambda p: {
            "isolated_cabg": "Isolated CABG", "single_non_cabg": "Single non-CABG",
            "two": "2 procedures", "three_plus": "3+ procedures",
        }.get(p.get("weight_of_procedure", "isolated_cabg"), "Isolated CABG"),
        lambda p: {"single_non_cabg": BETA["single_non_cabg"], "two": BETA["two"],
                    "three_plus": BETA["three_plus"]
                    }.get(p.get("weight_of_procedure", "isolated_cabg"), 0.0),
        "weight_of_procedure"),
    ("Surgery on thoracic aorta", lambda p: "Y" if p.get("thoracic_aorta") else "N",
        lambda p: BETA["thoracic_aorta"] if p.get("thoracic_aorta") else 0.0, "thoracic_aorta"),
]

SECTIONS = [
    ("Patient related factors", PATIENT_FIELDS),
    ("Cardiac related factors", CARDIAC_FIELDS),
    ("Operation related factors", OPERATION_FIELDS),
]

PATIENT_FILL = PatternFill("solid", fgColor="FCE4D6")
SECTION_FILL = PatternFill("solid", fgColor="DCE6F1")
SCORE_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fmt_beta(b):
    if b == 0:
        return "0"
    return f"{b:.7f}".rstrip("0").rstrip(".")


def write_block(ws, p, start_row):
    # Apply renal classification (same as cheatsheet)
    cc = cockcroft_gault(p.get("age"), p.get("weight_kg"),
                          p.get("cr_mg_dl"), p.get("female", False))
    if p.get("renal") == "dialysis":
        pass
    elif cc is not None:
        p["renal"] = renal_category(cc)
    score = calc(p)
    rationale = p.get("rationale", {}) or {}

    r = start_row
    # Title row spanning A:D
    cc_ref = f"CC={cc:.1f}" if cc is not None else "CC=?"
    title = (f"【{p['name']}】病歷號 {p['chart']}  |  "
             f"入院 {p.get('admit','?')} → {p.get('dc','?')}  |  "
             f"{p.get('age')}{'F' if p.get('female') else 'M'}  "
             f"{p.get('weight_kg')}kg  Cr {p.get('cr_mg_dl')}  {cc_ref}")
    cell = ws.cell(row=r, column=1, value=title)
    cell.font = Font(bold=True, size=12)
    cell.fill = PATIENT_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1

    # Column header
    headers = ["Field", "Value", "β", "Rationale / source"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = SECTION_FILL
        cell.border = BORDER
    r += 1

    for section_title, fields in SECTIONS:
        # section divider
        cell = ws.cell(row=r, column=1, value=section_title)
        cell.font = Font(bold=True, italic=True, color="555555")
        cell.fill = SECTION_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1
        for label, vf, bf, key in fields:
            ws.cell(row=r, column=1, value=label).border = BORDER
            ws.cell(row=r, column=2, value=vf(p)).border = BORDER
            ws.cell(row=r, column=3, value=fmt_beta(bf(p))).border = BORDER
            ws.cell(row=r, column=4, value=rationale.get(key, "")).border = BORDER
            r += 1

    # EuroSCORE II row
    ws.cell(row=r, column=1, value="EuroSCORE II").font = Font(bold=True, size=11)
    ws.cell(row=r, column=1).fill = SCORE_FILL
    score_cell = ws.cell(row=r, column=2, value=f"{score*100:.2f}%")
    score_cell.font = Font(bold=True, size=11)
    score_cell.fill = SCORE_FILL
    ws.cell(row=r, column=1).border = BORDER
    ws.cell(row=r, column=2).border = BORDER
    r += 2  # blank line between patients
    return r


def main():
    wb = openpyxl.load_workbook(PATH)
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 70

    yamls = sorted((ROOT / "_emr_raw").glob("*.yaml"))
    next_row = 1
    for yp in yamls:
        with open(yp, "r", encoding="utf-8") as f:
            p = yaml.safe_load(f)
        next_row = write_block(ws, p, next_row)

    wb.save(PATH)
    print(f"已新增 worksheet「{SHEET_NAME}」到 {PATH.name}，共 {len(yamls)} 位病人")
    print("欄位順序：MDCalc UI (Patient → Cardiac → Operation)")


if __name__ == "__main__":
    main()
