"""Create EURO -CR.xlsx with main sheet rows for our 4 patients.

Builds the workbook structure expected by generate_euroscore_sheet.py +
write_to_excel.py. After this runs, call:
    python generate_euroscore_sheet.py 陳常胤
    python write_to_excel.py
"""
import sys, io
from pathlib import Path
import openpyxl
import yaml

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).parent
PATH = ROOT / "EURO -CR.xlsx"
SHEET = "2026 (資訊室CAD-3.2)"
DOCTOR = "陳常胤"

# Column layout per generate_euroscore_sheet.py constants
HEADERS = [
    (1, "序"), (2, "入院"), (3, "出院"), (4, "病歷號"), (5, "姓名"),
    (6, "診斷"), (7, "VS"), (8, "CR均分"), (9, ""), (10, "EuroSCORE II %"),
]


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET
    for col, label in HEADERS:
        ws.cell(row=1, column=col, value=label)

    rows = []
    for ypath in sorted((ROOT / "_emr_raw").glob("*.yaml")):
        with open(ypath, "r", encoding="utf-8") as f:
            p = yaml.safe_load(f)
        rows.append(p)

    for i, p in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=p.get("admit"))
        ws.cell(row=i, column=3, value=p.get("dc"))
        ws.cell(row=i, column=4, value=p["chart"])
        ws.cell(row=i, column=5, value=p["name"])
        # diagnosis short hand
        dx_short = "CAD" if "ccs4" in p else ""
        ws.cell(row=i, column=6, value=dx_short)
        ws.cell(row=i, column=7, value="")  # VS attending (to be filled by user)
        ws.cell(row=i, column=8, value=DOCTOR)  # CR均分 — assigning to default doctor
        # col 9 empty, col 10 EURO score will be filled by write_to_excel.py
    wb.save(PATH)
    print(f"已建立 {PATH.name}，含 {len(rows)} 筆病人")
    for r in rows:
        print(f"  {r['chart']} {r['name']} {r.get('admit')} → {r.get('dc')}")


if __name__ == "__main__":
    main()
